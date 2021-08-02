import io
import re
import pandas as pd
import msoffcrypto
import json
from openpyxl import load_workbook
pd.set_option('display.width', None)

def find_check_result():
    f = pd.ExcelFile('C:\history\python\東京PCR検査記録表.xlsx')
    result = {}
    output_result = {}
    check_result = read_check_result()
    for name in f.sheet_names:
        check_list = []
        output_list = []
        output_dict = {}
        df = f.parse(sheet_name = name)
        df_list = list(df)
        for index, row in df[4:].iterrows():  # 根据数据模板，从第六行开始加载数据。
            if(str(row[1]).isdigit()):        # 根据数据模板，项目行名为:全半角数字
                for col_index in range(len(df_list)):          # 遍历获取每行中每一列的单元格值
                    if re.match(r'TOA', str(df.loc[index, df_list[col_index]])) and df.loc[index, df_list[col_index+2]] in ('L','H','再'):
                        check_list.append({df.loc[index, df_list[col_index]]:df.loc[index, df_list[col_index+2]]})
                        for check_item in check_result:
                            if check_item == str(df.loc[index, df_list[col_index]]):
                                # output_list.append({check_item:df.loc[index, df_list[col_index+2]]})
                                output_dict[check_item] = df.loc[index, df_list[col_index+2]]
                                output_result[name] = output_dict
            else:
                continue
        result[name] = check_list
    return output_result

def read_check_result():
    check_result_list = []
    file_path = 'C:\history\python\結果確認表.xlsx'

    of = pd.ExcelFile(file_path)
    for name in of.sheet_names:
        odf = of.parse(sheet_name = name)
        odf_list = list(odf)
        for index, row in odf[4:].iterrows():
            if(str(row[1]).isdigit()) and not pd.isnull(odf.loc[index, odf_list[4]]):
                check_result_list.append(odf.loc[index, odf_list[4]])  
                
    return check_result_list


def write_excel(output_write_result):
    output_file_path = 'C:\history\python\結果確認表.xlsx'
    wb = load_workbook(filename = output_file_path)
    print("output_write_result ",output_write_result)
    for name in wb.sheetnames:
        num_rows = 300       #从第七行，读取300行的'バーコード貼付欄'值
        num_columns = 5      #读取固定第六列的值
        print(name)
        for i in range (5, num_rows + 1):
            cell = wb[name].cell(row = i, column = num_columns)
            if re.match(r'TOA', str(cell.value)):
                # print(cell.value)
                for filename in output_write_result.keys():
                    print(filename)
                    print(output_write_result[filename].keys())
                    if cell.value in output_write_result[filename].keys():
                        wb[name].cell(row = i, column = num_columns+1).value = output_write_result[filename][cell.value]
                        wb[name].cell(row = i, column = num_columns+2).value = filename
    wb.save(output_file_path)

if __name__ == '__main__':
    output_write_result = find_check_result()
    write_excel(output_write_result)

