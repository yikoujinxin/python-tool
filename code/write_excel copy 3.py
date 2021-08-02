import xlrd
import msoffcrypto
import pandas as pd
import io
import re
from xlutils.copy import copy
from openpyxl import load_workbook
file_path = 'C:\history\python\要填结果表.xlsx'
wb = load_workbook(filename = file_path)
print("sheetnames ",wb.sheetnames)
for name in wb.sheetnames:
    num_rows = 4
    num_columns = 5
    for i in range (5, num_rows + 1):
        for j in range (6, num_columns + 1):
            cell = wb[name].cell(row = i, column = j)
            print(cell.value)
            if re.match(r'TOA', str(cell.value)):
                print(cell.value)