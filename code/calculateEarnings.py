import sys
import re
import pandas as pd
import json
from openpyxl import load_workbook,Workbook
pd.set_option('display.width', None)

#统计商品利润
def find_check_result(sheetName):
    f = pd.ExcelFile('.\商品別売上_20211101-20211101.xlsx')
    output_result = []
    check_result = read_check_result(sheetName) #统计商品成本
    df = f.parse(0) #使用第一个工作薄
    df_list = list(df)
    rowLength = 0 
    wb=Workbook()  #创建バーコードない記録工作表
    page = wb.active
    page.title = 'バーコードない記録'
    page.append(df_list)
    tempList = []
    
    for index, row in df[0:].iterrows():  # 从第二行开始加载数据。
        rowLength = rowLength + 1
        if(str(row[11]).isdigit()) and not pd.isnull(df.loc[index, df_list[11]]):  # 判断バーコード是否为数字，是否非null,nan
            for check_item in check_result: #判断バーコード是否相等
                if check_item[0] == str(row[11]):
                    print(index, row[0], row[11], check_item[1], row[3], row[7], int(row[3]) - (int(check_item[1]) * int(row[7])))
                    output_result.append((index, check_item[1], str(int(row[3]) - (int(check_item[1]) * int(row[7])))))
                    tempList.append(index)
    
    for index, row in df[0:].iterrows():  # 从第二行开始加载数据。
        if index not in tempList:
            page.append(row.tolist())
    wb.save('.\商品別売上_バーコードない記録.xlsx') 
    return output_result, rowLength

#统计商品成本
def read_check_result(sheetName):
    check_result_list = []
    file_path = '.\商品管理表_2021.11.8更新.xlsx'

    of = pd.ExcelFile(file_path)
    odf = of.parse(0) if sheetName == None else of.parse(sheet_name = sheetName)
    odf_list = list(odf)
    for index, row in odf[3:].iterrows():
        # print(index, str(row[5]), str(row[6]), pd.isnull(odf.loc[index, odf_list[10]]))
        if(str(row[6]).isdigit()) and not pd.isnull(odf.loc[index, odf_list[10]]): #判断バーコード是否成本和个数是否存在
            check_result_list.append((odf.loc[index, odf_list[6]], round(odf.loc[index, odf_list[10]])))
        
    return check_result_list

#将商品利润和商品总计记录下来
def write_excel(output_result, rowLength):
    output_file_path = r'.\tempFile.xlsx'
    wb = load_workbook(filename = output_file_path)
    print("output_result, rowLength ",output_result, rowLength)
    result, lastRow = output_result, rowLength
    total = 0
    for name in wb.sheetnames:
        print(name)
        for output in result:
            total = total + int(output[2])
            wb[name].cell(row = 1, column = 13).value = "仕入値（税込）"
            wb[name].cell(row = int(output[0])+2, column = 13).value = output[1]
            wb[name].cell(row = 1, column = 14).value = "売上"
            wb[name].cell(row = int(output[0])+2, column = 14).value = output[2]
            wb[name].cell(row = lastRow+2, column = 13).value = "合計"
            wb[name].cell(row = lastRow+2, column = 14).value = total
    wb.save(output_file_path)

if __name__ == '__main__':
    sheetName = sys.argv[1] if len(sys.argv)>1 else None #传入店铺名使用对应"店铺名"的工作薄，否则使用第一个工作薄
    output_result, rowLength = find_check_result(sheetName) #统计商品的成本和利润
    write_excel(output_result, rowLength) # 写入excel文件

