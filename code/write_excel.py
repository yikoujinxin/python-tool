import openpyxl
import xlrd
import msoffcrypto
import pandas as pd
import io
import re
from xlutils.copy import copy
from openpyxl import load_workbook
file_path = 'C:\history\python\結果対照表.xlsx'

def _load(file_name):
    return load_workbook(filename = file_name)

def _save(workbook, file_name):
    workbook.save(file_name)

pwb = _load(file_path)
# pwd.security.workbook_password = 'pcr-ht-n1'

wb = load_workbook(filename = file_path)

print("sheetnames ",wb.sheetnames)
for name in wb.sheetnames:
    num_rows = 100
    num_columns = 5
    print(name)
    for i in range (5, num_rows + 1):
        cell = wb[name].cell(row = i, column = num_columns)
        if re.match(r'TOA', str(cell.value)):
            print(cell.value)
            wb[name].cell(row = i, column = num_columns+1).value = 'L'
            wb[name].cell(row = i, column = num_columns+2).value = name
wb.save(file_path)

