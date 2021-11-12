import sys
import re
import json
import pandas as pd
import os

from openpyxl import Workbook, load_workbook


def first_step():
    try:
        sale = pd.read_csv("./商品別売上.csv", encoding="shift-jis")
        saleName_sample = pd.read_csv("./商品一括編集.csv", encoding="shift-jis", usecols=[4])
        janCode_sample = pd.read_csv("./商品一括編集.csv", encoding="shift-jis", usecols=[15])
    except IOError:
        print("Error:商品別売上.csvまたは商品一括編集.csvファイルを見つかりませんでした")
    sampleAll = pd.concat([saleName_sample, janCode_sample], axis=1)
    saleName = sale['商品名']
    saleNameList = saleName.values.tolist()
    sampleAllList = sampleAll.values.tolist()

    janCodeList = []
    k = 0

    for i in saleNameList:
        for j in sampleAllList:
            if i == j[0]:
                if k == 0:
                    if type(j[1]) == float:
                        continue
                    else:
                        j[1] = j[1].replace("#", "")
                        janCodeList.append(j[1])
                else:
                    continue
                k = 1
        if k == 1:
            k = 0
            continue
        else:
            janCodeList.append("null")
            k = 0
    sale['janCode'] = janCodeList
    sale.to_excel("統計済み.xlsx", index=None)


# 统计商品利润
def find_check_result(sheetName):
    try:
        f = pd.ExcelFile('./統計済み.xlsx')
    except IOError:
        print("Error:ファイル統計済み.xlsxは見つかりませんでした")
    output_result = []
    check_result = read_check_result(sheetName)  # 统计商品成本
    df = f.parse(0)  # 使用第一个工作薄
    df_list = list(df)
    rowLength = 0
    wb = Workbook()  # 创建バーコードない記録工作表
    page = wb.active
    page.title = 'バーコードない記録'
    page.append(df_list)
    tempList = []

    for index, row in df[0:].iterrows():  # 从第二行开始加载数据。
        rowLength = rowLength + 1
        if not pd.isnull(df.loc[index, df_list[11]]) and (str(round(row[11])).isdigit()):  # 判断バーコード是否为数字，是否非null,nan
            for check_item in check_result:  # 判断バーコード是否相等
                if check_item[0] == str(round(row[11])):
                    print(index, row[0], row[11], check_item[1], row[3], row[7],
                          int(row[3]) - (int(check_item[1]) * int(row[7])))
                    output_result.append((index, check_item[1], str(int(row[3]) - (int(check_item[1]) * int(row[7])))))
                    tempList.append(index)

    for index, row in df[0:].iterrows():  # 从第二行开始加载数据。
        if index not in tempList:
            page.append(row.tolist())
    wb.save('./商品別売上_バーコードない記録.xlsx')
    return output_result, rowLength


# 统计商品成本
def read_check_result(sheetName):
    check_result_list = []
    file_path = './商品管理表.xlsx'
    try:
        of = pd.ExcelFile(file_path)
    except IOError:
        print("ファイル商品管理表.xlsxは見つかりませんでした")
    odf = of.parse(0) if sheetName == None else of.parse(sheet_name=sheetName)
    odf_list = list(odf)
    for index, row in odf[3:].iterrows():
        # print(index, str(row[5]), str(row[6]), pd.isnull(odf.loc[index, odf_list[10]]))
        if (str(row[6]).isdigit()) and not pd.isnull(odf.loc[index, odf_list[10]]):  # 判断バーコード是否成本和个数是否存在
            check_result_list.append((odf.loc[index, odf_list[6]], round(odf.loc[index, odf_list[10]])))

    return check_result_list


# 将商品利润和商品总计记录下来
def write_excel(output_result, rowLength):
    output_file_path = r'.\統計済み.xlsx'
    wb = load_workbook(filename=output_file_path)
    print("output_result, rowLength ", output_result, rowLength)
    result, lastRow = output_result, rowLength
    total = 0
    for name in wb.sheetnames:
        print(name)
        for output in result:
            total = total + int(output[2])
            wb[name].cell(row=1, column=13).value = "仕入値（税込）"
            wb[name].cell(row=int(output[0]) + 2, column=13).value = output[1]
            wb[name].cell(row=1, column=14).value = "利潤"
            wb[name].cell(row=int(output[0]) + 2, column=14).value = output[2]
            wb[name].cell(row=lastRow + 2, column=13).value = "合計"
            wb[name].cell(row=lastRow + 2, column=14).value = total
    wb.save(output_file_path)


if __name__ == '__main__':
    first_step()
    # sheetName = sys.argv[1] if len(sys.argv) > 1 else None  # 传入店铺名使用对应"店铺名"的工作薄，否则使用第一个工作薄
    sheetName = input("商品管理表のシート名を入力してください:")
    output_result, rowLength = find_check_result(sheetName)  # 统计商品的成本和利润
    write_excel(output_result, rowLength)  # 写入excel文件
