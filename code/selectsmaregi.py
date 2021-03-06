import io
import re
import sys
import pandas as pd
import MySQLdb
# import msoffcrypto
# import json
from openpyxl import load_workbook
pd.set_option('display.width', None)

host="toa-cloud-test.cjrkfow6klcg.ap-northeast-1.rds.amazonaws.com"
user="shop_auction"
password="xjM2VxxIJZGHhlImnNt2yNntYGSLZBeG"
dbname="shop_auction"

def find_check_result():
    f = pd.ExcelFile('C:\history\python\東京PCR検査記録表.xlsx')
    result = {}
    output_result = {}
    check_result = read_check_result()
    for name in f.sheet_names:
        check_list = []
        output_list = []
        output_dict = {}
        df = f.parse(sheet_name = name)
        df_list = list(df)
        for index, row in df[4:].iterrows():  # 根据数据模板，从第六行开始加载数据。
            if(str(row[1]).isdigit()):        # 根据数据模板，项目行名为:全半角数字
                for col_index in range(len(df_list)):          # 遍历获取每行中每一列的单元格值
                    if re.match(r'TOA', str(df.loc[index, df_list[col_index]])) and df.loc[index, df_list[col_index+2]] in ('L','H','再'):
                        check_list.append({df.loc[index, df_list[col_index]]:df.loc[index, df_list[col_index+2]]})
                        for check_item in check_result:
                            if check_item == str(df.loc[index, df_list[col_index]]):
                                # output_list.append({check_item:df.loc[index, df_list[col_index+2]]})
                                output_dict[check_item] = df.loc[index, df_list[col_index+2]]
                                output_result[name] = output_dict
            else:
                continue
        result[name] = check_list
    return output_result

def read_check_result():
    check_result_list = []
    file_path = 'C:\history\python\結果確認表.xlsx'

    of = pd.ExcelFile(file_path)
    for name in of.sheet_names:
        odf = of.parse(sheet_name = name)
        odf_list = list(odf)
        for index, row in odf[4:].iterrows():
            if(str(row[1]).isdigit()) and not pd.isnull(odf.loc[index, odf_list[4]]):
                check_result_list.append(odf.loc[index, odf_list[4]])  
                
    return check_result_list

def read_good_names():
    keyNameArr = ["食品","肉","塩","醤油","オイスターソース","そば","飲","ミルク","キャンディ",
                    "餅","茶","コーン","酢","パイ","ゆであずき","菓","みそ","江崎グリコ","焼き",
                    "コンソメ","ドリンク","コカ","コーラ","糖","モンスターエナジー","カレー",
                    "豆乳","野菜","ポテト","炭酸水","サラダ","果汁","精米","カフェ","からし",
                    "あらしお","トマト","コーヒー","果実","ガーナリップル","ごま","スープ",
                    "パスタ","切餅","ラーメン","フィッシュ","わさび","おやつ","イソーセージ",
                    "カルパス","調味料","さば"]
    companyNameArr = ["明治","日清","アサヒ","ハウス","伊藤園","カルビー","SHINE","LIFE","味の素","CCL","カンロ",
                    "森永製菓","亀田製菓","大正製薬","大塚製薬","越後製菓","龍角散","ネスカフェ","湖池屋","S&B",
                    "ミツカン","ブルボン","永谷園","浪花屋","アイリスオーヤマ","キユーピー","サッポロ","カゴメ","サントリー",
                    "ハウス","ロッテ","カルピス","inバー","キリン","リスカ","キッコーマン","オリヒロ","マースジャパン"]
    check_food_list = []
    check_daily_goods_list = []
    deal_data=False
    file_path = 'C:\\pdf\\20211210_import可能.xlsx'

    of = pd.ExcelFile(file_path)
    for name in of.sheet_names:
        odf = of.parse(sheet_name = name)
        odf_list = list(odf)
        for index, row in odf[0:].iterrows():
            if any(item in str(row[3]) for item in keyNameArr):
                check_food_list.append(["150293",odf.loc[index, odf_list[2]],odf.loc[index, odf_list[3]],odf.loc[index, odf_list[1]]])
            elif any(item in str(row[3]) for item in companyNameArr):
                check_food_list.append(["150293",odf.loc[index, odf_list[2]],odf.loc[index, odf_list[3]],odf.loc[index, odf_list[1]]])
            else:
                check_food_list.append(["150311",odf.loc[index, odf_list[2]],odf.loc[index, odf_list[3]],odf.loc[index, odf_list[1]]]) 
    data_frame= pd.DataFrame(check_food_list)  
    data_frame.to_csv("C:\\pdf\\selectgoods.csv",index=False,sep=",",encoding='cp932')
    # return check_result_list

def insert_priority_smaregi_prices(file_path, db, cursor):
    of = pd.ExcelFile(file_path)
    odf = of.parse(sheet_name = "振分表②")
    odf_list = list(odf)
    for index, row in odf[0:1462].iterrows():
        if(row[13]) != 0 and not pd.isnull(odf.loc[index, odf_list[13]]):
            print(odf.loc[index, odf_list[6]],odf.loc[index, odf_list[8]],int(odf.loc[index, odf_list[14]]))
            insert_cmd = "insert into dv_clone_smaregi_tbl(commodity_name,commodity_jancode,smaregi_price,remark) values(\""
            insert_cmd = insert_cmd + str(odf.loc[index, odf_list[6]])+"\",\""+str(odf.loc[index, odf_list[8]])+"\",\""+str(odf.loc[index, odf_list[14]])+"\",\"優先価格を使用しました。\");"
            insert_cmd = insert_cmd
            cursor.execute(insert_cmd)
            db.commit()
def insert_smaregi_prices(file_path, db, cursor):
    insert_smaregi_list = []
    insert_zero_smaregi_list = []
    of = pd.ExcelFile(file_path)
    for name in of.sheet_names:
        odf = of.parse(sheet_name = name)
        odf_list = list(odf)
        for index, row in odf[0:].iterrows():
            select_clone_smaregi_cmd = "SELECT count(1) as cnt FROM dv_clone_smaregi_tbl WHERE commodity_jancode=\"" + str(row[3]) + "\";"
            cursor.execute(select_clone_smaregi_cmd)
            cursor.fetchall()
            for (cnt) in cursor:
                if int(cnt[0]) == 0:
                    if(row[6]) == 0 or pd.isnull(odf.loc[index, odf_list[6]]):
                        select_cmd = "SELECT summary_price FROM dv_summary_tbl WHERE commodity_jancode=\"" + str(row[3]) + "\";"
                        # print("select_cmd: ", select_cmd)
                        summary_price_count = cursor.execute(select_cmd)
                        summary_price_total = 0
                        cursor.fetchall()
                        for (summary_price) in cursor:
                            if summary_price[0] is not None:
                                summary_price_total += float(summary_price[0])
                            else:
                                print("Nonetype: ", summary_price[0])
                        if int(summary_price_count) > 0 and int(summary_price_total) > 0:
                            average_price = float(summary_price_total / summary_price_count)
                            print("summary_price: ", summary_price_total, summary_price_count, average_price, str(odf.loc[index, odf_list[3]]))
                            insert_zero_smaregi_list.append(["insert into dv_smaregi_tbl(commodity_name,commodity_jancode,smaregi_price,remark) values(\""+
                                str(odf.loc[index, odf_list[4]])+"\",\""+str(odf.loc[index, odf_list[3]])+"\",\""+
                                str(average_price)+"\",\"入力た価格を使用して、購入価格が見つかりませんでした\");"
                            ])
                        else:
                            insert_zero_smaregi_list.append(["insert into dv_smaregi_tbl(commodity_name,commodity_jancode,smaregi_price,remark) values(\""+
                                    str(odf.loc[index, odf_list[4]])+"\",\""+str(odf.loc[index, odf_list[3]])+"\",\""+
                                    str(odf.loc[index, odf_list[5]])+"\",\"販売価格の計算を使用して、購入価格が見つかりませんでした\");"
                                ])
                    else:
                        insert_smaregi_list.append(["insert into dv_smaregi_tbl(commodity_name,commodity_jancode,smaregi_price,remark) values(\""+
                                str(odf.loc[index, odf_list[4]])+"\",\""+str(odf.loc[index, odf_list[3]])+"\",\""+
                                str(odf.loc[index, odf_list[6]])+"\",\"\");"
                            ])
    data_frame2= pd.DataFrame(insert_zero_smaregi_list)  
    data_frame2.to_csv("C:\\pdf\\insert_zero_smaregi_list.csv",index=False,sep=",",encoding='cp932')

    data_frame= pd.DataFrame(insert_smaregi_list)  
    data_frame.to_csv("C:\\pdf\\insert_smaregi_sql.csv",index=False,sep=",",encoding='cp932')

def write_excel(output_write_result):
    output_file_path = 'C:\history\python\結果確認表.xlsx'
    wb = load_workbook(filename = output_file_path)
    print("output_write_result ",output_write_result)
    for name in wb.sheetnames:
        num_rows = 300       #从第七行，读取300行的'バーコード貼付欄'值
        num_columns = 5      #读取固定第六列的值
        print(name)
        for i in range (5, num_rows + 1):
            cell = wb[name].cell(row = i, column = num_columns)
            if re.match(r'TOA', str(cell.value)):
                # print(cell.value)
                for filename in output_write_result.keys():
                    print(filename)
                    print(output_write_result[filename].keys())
                    if cell.value in output_write_result[filename].keys():
                        wb[name].cell(row = i, column = num_columns+1).value = output_write_result[filename][cell.value]
                        wb[name].cell(row = i, column = num_columns+2).value = filename
    wb.save(output_file_path)

if __name__ == '__main__':
    file_path = sys.argv[1]
    # output_write_result = find_check_result()
    # write_excel(output_write_result)
    # read_good_names()
    db=MySQLdb.connect(host,user,password,dbname,charset="utf8")
    cursor=db.cursor()
    # priority_file = "C:\pdf\priority_file.xlsx"
    # insert_priority_smaregi_prices(priority_file, db, cursor)
    insert_smaregi_prices(file_path, db, cursor)
    

