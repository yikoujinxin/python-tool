import xlrd
import msoffcrypto
import pandas as pd
import io
import re
from xlutils.copy import copy
from openpyxl import load_workbook
file_path = 'C:\history\python\code\要填结果表.xlsx'

decrypted = io.BytesIO()
with open(file_path, "rb") as f:
    file = msoffcrypto.OfficeFile(f)
    file.load_key(password="pcr-ht-n1")  # Use password
    file.decrypt(decrypted)

of = pd.ExcelFile(decrypted)
print("sheetnames ",of.sheet_names)
for name in of.sheet_names:
    num_rows = 100
    num_columns = 5
    print(name)
    odf = of.parse(sheet_name = name)
    for i in range (5, num_rows + 1):
        cell = odf.iat[i, num_columns]
        if not pd.isnull(cell) and re.match(r'TOA', str(cell)):
            print(odf.iat[5,5])
            of.iat[ i, num_columns+1] = 'L'
            of.iat[ i, num_columns+2] = name

# wb = load_workbook(filename = file_path)
# print("sheetnames ",wb.sheetnames)
# for name in wb.sheetnames:
#     num_rows = 100
#     num_columns = 5
#     print(name)
#     for i in range (5, num_rows + 1):
#         cell = wb[name].cell(row = i, column = num_columns)
#         if re.match(r'TOA', str(cell.value)):
#             print(cell.value)
#             wb[name].cell(row = i, column = num_columns+1).value = 'L'
#             wb[name].cell(row = i, column = num_columns+2).value = name
of.save(file_path)