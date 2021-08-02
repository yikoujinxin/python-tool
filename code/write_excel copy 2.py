import xlrd
import tempfile
import msoffcrypto
import pandas as pd
import io
from openpyxl import load_workbook
file_path = 'C:\history\python\東京PCR検査記録表.xlsx'
decrypted = io.BytesIO()

with open(file_path, "rb") as f:
    file = msoffcrypto.OfficeFile(f)
    file.load_key(password="pcr-ht-n1")  # Use password
    file.decrypt(decrypted)

# of = pd.ExcelFile(decrypted)
fn = r'Temp.xlsx'
df = pd.read_excel(decrypted,header=None)
df2 = pd.DataFrame({'Data': [13, 24, 35, 46]})
writer = pd.ExcelWriter(file_path, engine='openpyxl')
book = load_workbook(decrypted)
writer.book = book

# writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

df.to_excel(writer, sheet_name='Sheet1', header=None, index=False)

df2.to_excel(writer, sheet_name='Sheet1', header=None, index=False,startcol=6,startrow=5)
writer.save()

# for name in of.sheet_names:
#     check_list = []
#     odf = of.parse(sheet_name = name)
#     odf_list = list(odf)
#     for index, row in odf[4:].iterrows():
#         if(str(row[1]).isdigit()) and not pd.isnull(odf.loc[index, odf_list[4]]):
#             print(odf.loc[index, odf_list[4]])



# file_path = 'C:\history\python\東京PCR検査記録表.xlsx'
# book = xlrd.open_workbook(file_path)
# sheet_1 = book.sheet_by_index(0)
# for col in range(sheet_1.ncols):
#     print('----------------------------')
#     for row in range(sheet_1.nrows):
#         print(sheet_1.cell(row, col).value)

